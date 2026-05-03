"""
Batch Prompt Tester — Batch Run 6
Tests OpenAI models via OpenRouter across prompt files.
Each prompt is run 5× per model at provider default temperature
and 1× per model at temperature=0.

Features:
    - robust RESUME from the flat CSV log using exact run keys
    - Excel workbook for browsing
    - flat CSV file for analysis (one row per model response)
    - JSON manifest with run settings
    - numeric answer parsing (raw + parsed output stored)
    - retry with exponential backoff + jitter
    - rate-limit aware global pauses
    - mostly sequential by default, with optional light concurrency

Requirements:
    pip install openai openpyxl

Setup:
    1. Sign up at openrouter.ai and get an API key
    2. Add credits at openrouter.ai/credits
    3. Set your key in the OPENROUTER_API_KEY environment variable
       or paste it below
    4. Run: python "batch run 6 updated.py"
"""

import csv
import json
import os
import random
import re
import threading
import time
from collections import deque
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from datetime import datetime, timezone
from pathlib import Path

from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────

OPENROUTER_API_KEY = ""

MODELS = [
    # (openrouter_model_id, short_label, company)
    ("openai/gpt-5.4",                       "GPT-5.4",              "OpenAI"),
    ("openai/gpt-5.4-mini",                  "GPT-5.4-mini",         "OpenAI"),
    ("google/gemini-2.5-pro",                "Gemini 2.5 Pro",       "Google"),
    ("google/gemini-2.5-flash",              "Gemini 2.5 Flash",     "Google"),
    ("google/gemini-3.1-pro-preview",        "Gemini 3.1 Pro",       "Google"),
    ("google/gemini-3.1-flash-lite-preview", "Gemini 3.1 Flash Lite","Google"),
    ("meta-llama/llama-4-maverick",          "Llama 4 Maverick",     "Meta"),
    ("meta-llama/llama-3.3-70b-instruct",    "Llama 3.3 70B",        "Meta"),
    ("mistralai/mistral-large",              "Mistral Large",        "Mistral"),
    ("mistralai/mistral-small-2603",         "Mistral Small",        "Mistral"),
    ("anthropic/claude-opus-4.6",            "Claude Opus 4.6",      "Anthropic"),
    ("anthropic/claude-sonnet-4.6",          "Claude Sonnet 4.6",    "Anthropic"),
    ("anthropic/claude-haiku-4.5",           "Claude Haiku 4.5",     "Anthropic"),
]

RUN_CONFIGS = [
    # (temperature, label_suffix, repeats)
    (None, "", 5),      # provider default temperature
    (0,    " T0", 1),   # deterministic baseline
]

MAX_TOKENS         = 2000
PROMPTS_FOLDER     = "prompts"
OUTPUT_FILE        = "results_batch_run_6.xlsx"
LOG_FILE           = "results_batch_run_6_raw.csv"
MANIFEST_FILE      = "results_batch_run_6_manifest.json"
DELAY_SECONDS      = 0.5
SAVE_EVERY         = 10
TRUNCATE_ERRORS    = 500
MAX_RETRIES        = 5
BACKOFF_BASE       = 2.0
BACKOFF_JITTER_MAX = 1.0

# Keep the script mostly sequential by default.
# You can raise this to 2 or 3 later if you want a very light level of concurrency.
MAX_WORKERS = 1

# Trigger a global pause if repeated likely rate-limit / overload errors occur.
RATE_LIMIT_WINDOW_SECONDS = 120
RATE_LIMIT_ERROR_THRESHOLD = 3
GLOBAL_PAUSE_SECONDS = 90

# Optional test mode: run only a small subset of prompts for quick validation.
TEST_MODE = False
TEST_PROMPT_LIMIT = 3

# Shuffle the full task list (model × temperature × repeat × prompt) to reduce time-based bias.
SHUFFLE_TASKS = True
SHUFFLE_SEED = 6

# ──────────────────────────────────────────────────────────────────────────────

HEADER_BLUE = "2E75B6"
ROW_ALT     = "F2F7FC"
ROW_WHITE   = "FFFFFF"

LANDLORD_LABELS = {
    'A': 'No framing',    'B': 'Comply+cheap',  'C': 'Cheap (illegal ok)',
    'D': 'Nice',          'E': 'Job at risk',   'F': 'Struggling financially',
}
PROBLEM_LABELS = {
    'A': 'Fridge (tenant)', 'B': 'Light bulb',  'C': 'Ink stain',
    'D': 'Laminate',        'E': 'Bathroom lock','F': 'Bad smell',
    'G': 'Toilet',          'H': 'Air con',      'I': 'Fridge (supplied)',
}
CONTEXT_LABELS = {'A': 'None', 'B': 'Sympathetic', 'C': 'Threatening'}
LEGAL_LABELS   = {
    'A': 'No law',  'B': 'Excerpt',  'C': 'Full schedule',
    'D': 'Sched+relevant cases',  'E': 'Sched+all cases',
}

VALID_ANSWERS = {"1", "2", "3"}

CSV_HEADERS = [
    "run_key",
    "prompt_id",
    "filename",
    "model",
    "model_id",
    "company",
    "temperature",
    "temperature_label",
    "sheet_name",
    "repeat",
    "landlord",
    "legal",
    "problem",
    "context",
    "raw_response",
    "parsed_response",
    "parse_status",
    "format_followed",
    "response",
    "error",
    "latency_seconds",
    "prompt_tokens",
    "completion_tokens",
    "total_tokens",
    "finish_reason",
    "attempts",
    "completed_at_utc",
]

MODEL_SHEET_HEADERS = [
    "Run Key",
    "Prompt ID",
    "Filename",
    "Landlord",
    "Legal",
    "Problem",
    "Context",
    "Repeat",
    "Raw response",
    "Parsed response",
    "Parse status",
    "Format followed",
    "Error",
    "Latency (s)",
    "Prompt tokens",
    "Completion tokens",
    "Total tokens",
    "Finish reason",
    "Attempts",
    "Completed at UTC",
]
MODEL_SHEET_WIDTHS = [22, 40, 52, 22, 26, 20, 14, 10, 28, 16, 14, 14, 30, 12, 14, 16, 12, 14, 10, 22]


class GlobalPauseController:
    def __init__(self, threshold, window_seconds, pause_seconds):
        self.threshold = threshold
        self.window_seconds = window_seconds
        self.pause_seconds = pause_seconds
        self.error_times = deque()
        self.pause_until = 0.0
        self.lock = threading.Lock()

    def wait_if_paused(self):
        while True:
            with self.lock:
                remaining = self.pause_until - time.time()
            if remaining <= 0:
                return
            time.sleep(min(remaining, 1.0))

    def record_error(self, error_text):
        if not is_rate_limit_like_error(error_text):
            return False

        now = time.time()
        with self.lock:
            self.error_times.append(now)
            cutoff = now - self.window_seconds
            while self.error_times and self.error_times[0] < cutoff:
                self.error_times.popleft()

            if len(self.error_times) >= self.threshold:
                new_pause_until = now + self.pause_seconds
                if new_pause_until > self.pause_until:
                    self.pause_until = new_pause_until
                    return True
        return False


def parse_filename(filename):
    stem = Path(filename).stem
    result = {}
    for part in stem.split('_'):
        if part.startswith('Landlord'):
            result['landlord'] = part[8:]
        elif part.startswith('Legal'):
            result['legal'] = part[5:]
        elif part.startswith('Problem'):
            result['problem'] = part[7:]
        elif part.startswith('Context'):
            result['context'] = part[7:]
        elif part and part[0].isdigit():
            result['problem_name'] = part
    return result


def get_prompt_id(filename):
    return Path(filename).stem


def format_temperature_label(temperature):
    return "default" if temperature is None else str(temperature)


def build_run_key(prompt_id, model_id, temperature, repeat):
    return "|".join([prompt_id, model_id, format_temperature_label(temperature), str(repeat)])


def style_header(ws, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, name="Arial", color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=HEADER_BLUE)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def style_row(ws, row, n_cols):
    fill = ROW_ALT if row % 2 == 0 else ROW_WHITE
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", start_color=fill)
        cell.font = Font(name="Arial")
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def utc_now_iso():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def parse_numeric_answer(text):
    if not text:
        return "", "empty", False

    cleaned = str(text).strip()

    if cleaned in VALID_ANSWERS:
        return cleaned, "exact", True

    matches = [m for m in re.findall(r"\b\d+\b", cleaned) if m in VALID_ANSWERS]
    unique_matches = list(dict.fromkeys(matches))

    if len(unique_matches) == 1:
        return unique_matches[0], "extracted", False
    if len(unique_matches) > 1:
        return "", "ambiguous", False
    return "", "invalid", False


def is_rate_limit_like_error(error_text):
    if not error_text:
        return False
    text = str(error_text).lower()
    indicators = [
        "429",
        "rate limit",
        "too many requests",
        "overloaded",
        "capacity",
        "try again later",
        "temporarily unavailable",
        "server overloaded",
        "throttle",
        "quota exceeded",
        "provider returned error",
        "timeout",
        "timed out",
    ]
    return any(indicator in text for indicator in indicators)


def make_client(api_key):
    return OpenAI(base_url="https://openrouter.ai/api/v1", api_key=api_key)


def migrate_log_file_if_needed():
    path = Path(LOG_FILE)
    if not path.exists() or path.stat().st_size == 0:
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(CSV_HEADERS)
        return

    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        existing_headers = reader.fieldnames or []
        rows = list(reader)

    if existing_headers == CSV_HEADERS:
        return

    existing_header_set = set(existing_headers)
    required_legacy_subset = {
        "prompt_id", "filename", "model", "model_id", "company",
        "temperature", "temperature_label", "sheet_name", "repeat",
        "landlord", "legal", "problem", "context", "response", "error"
    }
    if not required_legacy_subset.issubset(existing_header_set):
        raise ValueError(
            f"Existing CSV '{LOG_FILE}' has an incompatible header. "
            "Please back it up or rename it before running this upgraded script."
        )

    migrated_rows = []
    for row in rows:
        prompt_id = row.get("prompt_id", "")
        model_id = row.get("model_id", "")
        repeat = row.get("repeat", "")
        temp_label = row.get("temperature_label", "default")
        temperature_raw = row.get("temperature", "")
        temperature = None if temp_label == "default" else temperature_raw
        run_key = row.get("run_key") or build_run_key(prompt_id, model_id, temperature, repeat)

        legacy_response = row.get("response", "")
        parsed_response, parse_status, format_followed = parse_numeric_answer(legacy_response)

        migrated_rows.append({
            "run_key": run_key,
            "prompt_id": prompt_id,
            "filename": row.get("filename", ""),
            "model": row.get("model", ""),
            "model_id": model_id,
            "company": row.get("company", ""),
            "temperature": row.get("temperature", ""),
            "temperature_label": temp_label,
            "sheet_name": row.get("sheet_name", ""),
            "repeat": row.get("repeat", ""),
            "landlord": row.get("landlord", ""),
            "legal": row.get("legal", ""),
            "problem": row.get("problem", ""),
            "context": row.get("context", ""),
            "raw_response": row.get("raw_response", legacy_response),
            "parsed_response": row.get("parsed_response", parsed_response),
            "parse_status": row.get("parse_status", parse_status),
            "format_followed": row.get("format_followed", format_followed),
            "response": row.get("parsed_response", parsed_response) or legacy_response,
            "error": row.get("error", ""),
            "latency_seconds": row.get("latency_seconds", ""),
            "prompt_tokens": row.get("prompt_tokens", ""),
            "completion_tokens": row.get("completion_tokens", ""),
            "total_tokens": row.get("total_tokens", ""),
            "finish_reason": row.get("finish_reason", ""),
            "attempts": row.get("attempts", ""),
            "completed_at_utc": row.get("completed_at_utc", ""),
        })

    backup = path.with_suffix(path.suffix + ".bak")
    path.replace(backup)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(migrated_rows)

    print(f"Migrated existing CSV to the new schema. Backup saved as: {backup.name}")


def append_csv_row(row):
    with open(LOG_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writerow(row)


def load_existing_csv_state(valid_sheet_names, valid_prompt_filenames):
    completed_keys = set()
    response_lookup = {filename: {sheet_name: "" for sheet_name in valid_sheet_names} for filename in valid_prompt_filenames}
    csv_rows = []

    path = Path(LOG_FILE)
    if not path.exists() or path.stat().st_size == 0:
        return completed_keys, response_lookup, csv_rows

    valid_prompt_filenames = set(valid_prompt_filenames)
    valid_sheet_names = set(valid_sheet_names)

    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            run_key = row.get("run_key", "")
            filename = row.get("filename", "")
            sheet_name = row.get("sheet_name", "")
            response_text = row.get("parsed_response", row.get("response", ""))

            if run_key:
                completed_keys.add(run_key)

            if filename in valid_prompt_filenames and sheet_name in valid_sheet_names:
                response_lookup[filename][sheet_name] = response_text
                csv_rows.append(row)

    return completed_keys, response_lookup, csv_rows


def ensure_model_sheet(ws):
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        for col, h in enumerate(MODEL_SHEET_HEADERS, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, len(MODEL_SHEET_HEADERS))
        set_col_widths(ws, MODEL_SHEET_WIDTHS)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(MODEL_SHEET_HEADERS))}1"


def ensure_sheet_exists(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ensure_model_sheet(ws)
        return ws

    ws = wb[sheet_name]
    if ws.max_row < 1:
        ensure_model_sheet(ws)
    elif ws.cell(row=1, column=1).value != MODEL_SHEET_HEADERS[0]:
        if ws.max_row == 1 and all(ws.cell(row=1, column=i).value in (None, "") for i in range(1, len(MODEL_SHEET_HEADERS) + 1)):
            ensure_model_sheet(ws)
    return ws


def worksheet_run_keys(ws):
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        run_key = row[0] if row else None
        if run_key:
            keys.add(str(run_key))
    return keys


def append_model_sheet_row(ws, row_dict):
    next_row = ws.max_row + 1
    values = [
        row_dict["run_key"],
        row_dict["prompt_id"],
        row_dict["filename"],
        row_dict["landlord"],
        row_dict["legal"],
        row_dict["problem"],
        row_dict["context"],
        row_dict["repeat"],
        row_dict["raw_response"],
        row_dict["parsed_response"],
        row_dict["parse_status"],
        row_dict["format_followed"],
        row_dict["error"],
        row_dict["latency_seconds"],
        row_dict["prompt_tokens"],
        row_dict["completion_tokens"],
        row_dict["total_tokens"],
        row_dict["finish_reason"],
        row_dict["attempts"],
        row_dict["completed_at_utc"],
    ]
    for col, value in enumerate(values, 1):
        ws.cell(row=next_row, column=col, value=value)
    style_row(ws, next_row, len(MODEL_SHEET_HEADERS))


def sync_workbook_from_csv(wb, csv_rows, valid_sheet_names):
    valid_sheet_names = set(valid_sheet_names)
    worksheet_keys = {}
    appended = 0

    for row in csv_rows:
        sheet_name = row.get("sheet_name", "")
        run_key = row.get("run_key", "")
        if not sheet_name or not run_key or sheet_name not in valid_sheet_names:
            continue

        ws = ensure_sheet_exists(wb, sheet_name)
        if sheet_name not in worksheet_keys:
            worksheet_keys[sheet_name] = worksheet_run_keys(ws)

        if run_key in worksheet_keys[sheet_name]:
            continue

        append_model_sheet_row(ws, row)
        worksheet_keys[sheet_name].add(run_key)
        appended += 1

    return appended


def call_api(client, model_id, prompt_text, temperature=None, pause_controller=None):
    kwargs = {
        "model": model_id,
        "max_tokens": MAX_TOKENS,
        "messages": [{"role": "user", "content": prompt_text}],
        "extra_headers": {
            "HTTP-Referer": "https://legal-alignment-experiment.research",
            "X-Title": "Legal Alignment Experiment Batch Run 6",
        },
        "extra_body": {"reasoning": {"max_tokens": 2000}},
    }
    if temperature is not None:
        kwargs["temperature"] = temperature

    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        if pause_controller:
            pause_controller.wait_if_paused()

        started = time.perf_counter()
        try:
            response = client.chat.completions.create(**kwargs)
            latency_seconds = round(time.perf_counter() - started, 3)

            content = response.choices[0].message.content
            usage = getattr(response, "usage", None)
            choice = response.choices[0]

            return {
                "response_text": content.strip() if content else "",
                "error_text": "",
                "latency_seconds": latency_seconds,
                "prompt_tokens": getattr(usage, "prompt_tokens", "") if usage else "",
                "completion_tokens": getattr(usage, "completion_tokens", "") if usage else "",
                "total_tokens": getattr(usage, "total_tokens", "") if usage else "",
                "finish_reason": getattr(choice, "finish_reason", "") or "",
                "attempts": attempt,
            }
        except Exception as e:
            last_error = e
            error_text = str(e)[:TRUNCATE_ERRORS]

            triggered_pause = False
            if pause_controller:
                triggered_pause = pause_controller.record_error(error_text)

            if triggered_pause:
                print(f"\n[global pause] Repeated rate-limit/overload errors detected. Pausing for {GLOBAL_PAUSE_SECONDS}s.\n", flush=True)

            if attempt == MAX_RETRIES:
                break

            sleep_time = (BACKOFF_BASE ** (attempt - 1)) + random.uniform(0, BACKOFF_JITTER_MAX)
            print(f"retrying in {sleep_time:.2f}s", end=" ... ", flush=True)
            time.sleep(sleep_time)

    return {
        "response_text": "",
        "error_text": str(last_error)[:TRUNCATE_ERRORS] if last_error else "Unknown error",
        "latency_seconds": "",
        "prompt_tokens": "",
        "completion_tokens": "",
        "total_tokens": "",
        "finish_reason": "",
        "attempts": MAX_RETRIES,
    }


def build_csv_row(prompt_file, parsed, label, model_id, company, temperature, sheet_name, repeat, result):
    prompt_id = get_prompt_id(prompt_file.name)
    raw_response = result["response_text"]
    parsed_response, parse_status, format_followed = parse_numeric_answer(raw_response)
    return {
        "run_key": build_run_key(prompt_id, model_id, temperature, repeat),
        "prompt_id": prompt_id,
        "filename": prompt_file.name,
        "model": label,
        "model_id": model_id,
        "company": company,
        "temperature": "" if temperature is None else temperature,
        "temperature_label": format_temperature_label(temperature),
        "sheet_name": sheet_name,
        "repeat": repeat,
        "landlord": LANDLORD_LABELS.get(parsed.get('landlord', ''), ''),
        "legal": LEGAL_LABELS.get(parsed.get('legal', ''), ''),
        "problem": PROBLEM_LABELS.get(parsed.get('problem', ''), parsed.get('problem_name', '')),
        "context": CONTEXT_LABELS.get(parsed.get('context', ''), ''),
        "raw_response": raw_response,
        "parsed_response": parsed_response,
        "parse_status": parse_status,
        "format_followed": format_followed,
        "response": parsed_response,
        "error": result["error_text"],
        "latency_seconds": result["latency_seconds"],
        "prompt_tokens": result["prompt_tokens"],
        "completion_tokens": result["completion_tokens"],
        "total_tokens": result["total_tokens"],
        "finish_reason": result["finish_reason"],
        "attempts": result["attempts"],
        "completed_at_utc": utc_now_iso(),
    }


def task_run_key(task):
    model_id, _label, _company, temperature, _sheet_name, repeat, prompt_file = task
    return build_run_key(get_prompt_id(prompt_file.name), model_id, temperature, repeat)


def build_task_list(runs, prompt_files):
    tasks = []
    for model_id, label, company, temperature, sheet_name, repeat in runs:
        for prompt_file in prompt_files:
            tasks.append((model_id, label, company, temperature, sheet_name, repeat, prompt_file))
    return tasks


def write_manifest(prompt_files, runs):
    manifest = {
        "run_name": "batch run 6",
        "script_name": Path(__file__).name if "__file__" in globals() else "batch run 6 updated.py",
        "created_at_utc": utc_now_iso(),
        "prompt_folder": PROMPTS_FOLDER,
        "prompt_count": len(prompt_files),
        "output_file": OUTPUT_FILE,
        "log_file": LOG_FILE,
        "manifest_file": MANIFEST_FILE,
        "models": [
            {"model_id": model_id, "label": label, "company": company}
            for model_id, label, company in MODELS
        ],
        "run_configs": [
            {
                "temperature": temperature,
                "temperature_label": format_temperature_label(temperature),
                "label_suffix": suffix,
                "repeats": repeats,
            }
            for temperature, suffix, repeats in RUN_CONFIGS
        ],
        "max_tokens": MAX_TOKENS,
        "valid_answers": sorted(VALID_ANSWERS),
        "response_parsing": {
            "strategy": "extract a single numeric answer from raw output",
            "stores_raw_response": True,
            "stores_parsed_response": True,
            "stores_format_followed": True,
        },
        "delay_seconds": DELAY_SECONDS,
        "max_retries": MAX_RETRIES,
        "backoff_base": BACKOFF_BASE,
        "backoff_jitter_max": BACKOFF_JITTER_MAX,
        "max_workers": MAX_WORKERS,
        "rate_limit_window_seconds": RATE_LIMIT_WINDOW_SECONDS,
        "rate_limit_error_threshold": RATE_LIMIT_ERROR_THRESHOLD,
        "global_pause_seconds": GLOBAL_PAUSE_SECONDS,
        "test_mode": TEST_MODE,
        "test_prompt_limit": TEST_PROMPT_LIMIT,
        "shuffle_tasks": SHUFFLE_TASKS,
        "shuffle_seed": SHUFFLE_SEED,
        "expected_api_calls": len(prompt_files) * len(runs),
        "run_variants": [
            {
                "model_id": model_id,
                "label": label,
                "company": company,
                "temperature": temperature,
                "temperature_label": format_temperature_label(temperature),
                "sheet_name": sheet_name,
                "repeat": repeat,
            }
            for model_id, label, company, temperature, sheet_name, repeat in runs
        ],
    }

    with open(MANIFEST_FILE, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)


def run_single_prompt(prompt_file, model_id, label, company, temperature, sheet_name, repeat, api_key, pause_controller):
    client = make_client(api_key)
    parsed = parse_filename(prompt_file.name)
    prompt_text = prompt_file.read_text(encoding="utf-8")
    result = call_api(client, model_id, prompt_text, temperature, pause_controller=pause_controller)
    row_dict = build_csv_row(
        prompt_file=prompt_file,
        parsed=parsed,
        label=label,
        model_id=model_id,
        company=company,
        temperature=temperature,
        sheet_name=sheet_name,
        repeat=repeat,
        result=result,
    )
    return row_dict


def print_result_summary(row_dict):
    if row_dict["error"]:
        print(f"ERROR after {row_dict['attempts']} attempt(s): {row_dict['error'][:80]}")
    else:
        print(
            f"-> raw={row_dict['raw_response'][:40]!r} "
            f"parsed={row_dict['parsed_response']!r} "
            f"status={row_dict['parse_status']} "
            f"format_followed={row_dict['format_followed']}"
        )


def run_batch():
    api_key = (
        OPENROUTER_API_KEY
        if OPENROUTER_API_KEY != "YOUR_OPENROUTER_KEY_HERE"
        else os.environ.get("OPENROUTER_API_KEY")
    )
    if not api_key:
        raise ValueError("Please set your OpenRouter API key in the script or environment.")

    if MAX_WORKERS < 1:
        raise ValueError("MAX_WORKERS must be at least 1.")
    if MAX_WORKERS > 3:
        raise ValueError("MAX_WORKERS is capped at 3 for this script.")

    prompt_files = sorted(Path(PROMPTS_FOLDER).glob("*.txt"))
    if not prompt_files:
        raise FileNotFoundError(f"No .txt files found in '{PROMPTS_FOLDER}'")

    full_prompt_count = len(prompt_files)
    if TEST_MODE:
        if TEST_PROMPT_LIMIT < 1:
            raise ValueError("TEST_PROMPT_LIMIT must be at least 1 when TEST_MODE is enabled.")
        prompt_files = prompt_files[:TEST_PROMPT_LIMIT]

    migrate_log_file_if_needed()

    runs = []
    for model_id, label, company in MODELS:
        for temp, suffix, repeats in RUN_CONFIGS:
            for r in range(repeats):
                run_label = f"{label}{suffix} R{r+1}" if repeats > 1 else f"{label}{suffix}"
                runs.append((model_id, label, company, temp, run_label, r + 1))

    total = len(prompt_files)
    total_calls = total * len(runs)
    all_sheet_labels = [sheet_name for *_, sheet_name, _ in runs]
    all_prompt_filenames = [pf.name for pf in prompt_files]

    tasks = build_task_list(runs, prompt_files)
    if SHUFFLE_TASKS:
        random.seed(SHUFFLE_SEED)
        random.shuffle(tasks)

    write_manifest(prompt_files, runs)

    if TEST_MODE:
        print(f"TEST MODE ENABLED: using {total} of {full_prompt_count} prompt files.")
    else:
        print(f"Found {total} prompt files.")
    print(f"{len(MODELS)} models, {sum(repeats for _, _, repeats in RUN_CONFIGS)} runs each = {len(runs)} run variants")
    print(f"Total API calls if starting fresh: {total_calls:,}")
    print(f"Task order: {'full shuffled' if SHUFFLE_TASKS else 'grouped/unshuffled'}")
    if SHUFFLE_TASKS:
        print(f"Shuffle seed:   {SHUFFLE_SEED}")
    print(f"Execution mode: {'sequential' if MAX_WORKERS == 1 else f'light concurrency ({MAX_WORKERS} workers)'}")
    print(f"Excel output:   {OUTPUT_FILE}")
    print(f"CSV log:        {LOG_FILE}")
    print(f"Run manifest:   {MANIFEST_FILE}\n")

    completed_keys, all_responses, existing_csv_rows = load_existing_csv_state(
        valid_sheet_names=all_sheet_labels,
        valid_prompt_filenames=all_prompt_filenames,
    )

    print(f"Existing completed runs in CSV: {len(completed_keys):,}\n")

    if Path(OUTPUT_FILE).exists():
        print(f"Found existing {OUTPUT_FILE} - loading workbook and syncing from CSV...\n")
        wb = load_workbook(OUTPUT_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    synced_rows = sync_workbook_from_csv(wb, existing_csv_rows, all_sheet_labels)
    if synced_rows:
        wb.save(OUTPUT_FILE)
        print(f"Synced {synced_rows} missing workbook rows from the CSV log.\n")

    pause_controller = GlobalPauseController(
        threshold=RATE_LIMIT_ERROR_THRESHOLD,
        window_seconds=RATE_LIMIT_WINDOW_SECONDS,
        pause_seconds=GLOBAL_PAUSE_SECONDS,
    )

    remaining_tasks = [task for task in tasks if task_run_key(task) not in completed_keys]
    print(f"Remaining API calls to run: {len(remaining_tasks):,}\n")

    if not remaining_tasks:
        print("No remaining tasks. Skipping straight to comparison sheet rebuild.")
    save_counter = 0
    sheet_cache = {}

    def get_ws(sheet_name):
        if sheet_name not in sheet_cache:
            sheet_cache[sheet_name] = ensure_sheet_exists(wb, sheet_name)
        return sheet_cache[sheet_name]

    if MAX_WORKERS == 1:
        for idx, task in enumerate(remaining_tasks, 1):
            model_id, label, company, temperature, sheet_name, repeat, prompt_file = task
            temp_str = f"temp={temperature}" if temperature is not None else "temp=default"
            print(
                f"[{idx}/{len(remaining_tasks)}] {company} / {sheet_name} "
                f"({model_id}, {temp_str}, repeat {repeat}) | {prompt_file.name} ... ",
                end="",
                flush=True,
            )
            row_dict = run_single_prompt(
                prompt_file=prompt_file,
                model_id=model_id,
                label=label,
                company=company,
                temperature=temperature,
                sheet_name=sheet_name,
                repeat=repeat,
                api_key=api_key,
                pause_controller=pause_controller,
            )
            print_result_summary(row_dict)

            append_csv_row(row_dict)
            append_model_sheet_row(get_ws(sheet_name), row_dict)

            completed_keys.add(row_dict["run_key"])
            all_responses[row_dict["filename"]][sheet_name] = row_dict["parsed_response"]
            save_counter += 1

            if save_counter % SAVE_EVERY == 0:
                wb.save(OUTPUT_FILE)

            time.sleep(DELAY_SECONDS)
    else:
        worker_count = max(1, min(MAX_WORKERS, 3))
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            pending = {}
            remaining_iter = iter(remaining_tasks)

            def submit_next():
                try:
                    task = next(remaining_iter)
                except StopIteration:
                    return False
                model_id, label, company, temperature, sheet_name, repeat, prompt_file = task
                future = executor.submit(
                    run_single_prompt,
                    prompt_file,
                    model_id,
                    label,
                    company,
                    temperature,
                    sheet_name,
                    repeat,
                    api_key,
                    pause_controller,
                )
                pending[future] = task
                return True

            for _ in range(min(worker_count, len(remaining_tasks))):
                submit_next()

            done_count = 0
            while pending:
                done, _ = wait(list(pending.keys()), return_when=FIRST_COMPLETED)
                for future in done:
                    task = pending.pop(future)
                    model_id, _label, company, temperature, sheet_name, repeat, prompt_file = task
                    done_count += 1
                    temp_str = f"temp={temperature}" if temperature is not None else "temp=default"
                    print(
                        f"[{done_count}/{len(remaining_tasks)}] {company} / {sheet_name} "
                        f"({model_id}, {temp_str}, repeat {repeat}) | {prompt_file.name} ... ",
                        end="",
                        flush=True,
                    )

                    row_dict = future.result()
                    print_result_summary(row_dict)

                    append_csv_row(row_dict)
                    append_model_sheet_row(get_ws(sheet_name), row_dict)

                    completed_keys.add(row_dict["run_key"])
                    all_responses[row_dict["filename"]][sheet_name] = row_dict["parsed_response"]
                    save_counter += 1

                    if save_counter % SAVE_EVERY == 0:
                        wb.save(OUTPUT_FILE)

                    time.sleep(DELAY_SECONDS)
                    submit_next()

    wb.save(OUTPUT_FILE)

    print("\nBuilding comparison sheet...")
    if "Comparison" in wb.sheetnames:
        del wb["Comparison"]

    comp_headers = ["Prompt ID", "Filename", "Landlord", "Legal", "Problem", "Context"] + all_sheet_labels + ["All agree?"]
    comp_widths = [40, 52, 22, 26, 20, 14] + [16] * len(all_sheet_labels) + [12]
    agree_col = len(comp_headers)

    ws_comp = wb.create_sheet(title="Comparison", index=0)
    for col, h in enumerate(comp_headers, 1):
        ws_comp.cell(row=1, column=col, value=h)
    style_header(ws_comp, len(comp_headers))
    set_col_widths(ws_comp, comp_widths)
    ws_comp.freeze_panes = "A2"
    ws_comp.auto_filter.ref = f"A1:{get_column_letter(len(comp_headers))}1"

    for i, prompt_file in enumerate(prompt_files, 1):
        parsed = parse_filename(prompt_file.name)
        prompt_id = get_prompt_id(prompt_file.name)
        responses = all_responses[prompt_file.name]
        row = i + 1

        ws_comp.cell(row=row, column=1, value=prompt_id)
        ws_comp.cell(row=row, column=2, value=prompt_file.name)
        ws_comp.cell(row=row, column=3, value=LANDLORD_LABELS.get(parsed.get('landlord', ''), ''))
        ws_comp.cell(row=row, column=4, value=LEGAL_LABELS.get(parsed.get('legal', ''), ''))
        ws_comp.cell(row=row, column=5, value=PROBLEM_LABELS.get(parsed.get('problem', ''), parsed.get('problem_name', '')))
        ws_comp.cell(row=row, column=6, value=CONTEXT_LABELS.get(parsed.get('context', ''), ''))

        vals = []
        for col_offset, sl in enumerate(all_sheet_labels):
            response = responses.get(sl, "")
            ws_comp.cell(row=row, column=7 + col_offset, value=response)
            vals.append(response)

        nonempty_vals = [str(v).strip().lower() for v in vals if v and str(v).strip()]
        all_agree = len(set(nonempty_vals)) == 1 if nonempty_vals else False
        agree_cell = ws_comp.cell(row=row, column=agree_col, value="Yes" if all_agree else "No")
        if not all_agree:
            agree_cell.font = Font(name="Arial", bold=True, color="C00000")
        style_row(ws_comp, row, len(comp_headers))

    wb.save(OUTPUT_FILE)
    print(f"\nDone! Results saved to: {OUTPUT_FILE}")
    print(f"Flat CSV log saved to: {LOG_FILE}")
    print(f"Run manifest saved to: {MANIFEST_FILE}")
    print(f"Sheets: Comparison + {len(runs)} model/temp/repeat variants")
    print(f"Total prompts: {len(prompt_files)}")


if __name__ == "__main__":
    run_batch()
